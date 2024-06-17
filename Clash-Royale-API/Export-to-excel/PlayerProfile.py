import openpyxl
import requests
import time
from tqdm import tqdm

#Calculate starting time
start_time = time.time()

# Enter your key
# Go to https://developer.clashroyale.com/#/ 

API_KEY = ""
headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

# Enter your tag
# ex: '%2322R920J00','%2312345678'

player_tags = ['%2322R920J00']
              
wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "Tag"
ws["B1"] = "Classic12Wins"
ws["C1"] = "Grand12Wins"
ws["D1"] = "YearsPlayed"
ws["E1"] = "EmoteCollection"
ws["F1"] = "BannerCollection"
ws["G1"] = "starPoints"
ws["H1"] = "totalExpPoints"


for player_tag in tqdm(player_tags):
    response = requests.get(f"https://api.clashroyale.com/v1/players/{player_tag}", headers=headers)
    
    player_data = response.json()

    Classic12WinsCout =0 
    Grand12WinsCout = 0

    starPoints = player_data["starPoints"]
    totalExpPoints = player_data["totalExpPoints"]
    
    for Classic12Wins in player_data["badges"]:
       if Classic12Wins["name"] == "Classic12Wins":
        Classic12WinsCout = Classic12Wins["progress"]

    for Grand12Wins in player_data["badges"]:
       if Grand12Wins["name"] == "Grand12Wins":
        Grand12WinsCout = Grand12Wins["progress"]

    for YearsPlayed in player_data["badges"]:
       if YearsPlayed["name"] == "YearsPlayed":
        YearsPlayedCout = YearsPlayed["progress"]

    for EmoteCollection in player_data["badges"]:
       if EmoteCollection["name"] == "EmoteCollection":
        EmoteCollectionCout = EmoteCollection["progress"]

    for BannerCollection in player_data["badges"]:
       if BannerCollection["name"] == "BannerCollection":
        BannerCollectionCout = BannerCollection["progress"]


    # for PracticewithFriendsCout in player_data["achievements"]:
    #   if PracticewithFriendsCout["name"] == "Practice with Friends":
    #     PracticewithFriendsCoutCout = PracticewithFriendsCout["value"]

    ws.append([
    player_data["name"],Classic12WinsCout,Grand12WinsCout,YearsPlayedCout,EmoteCollectionCout,BannerCollectionCout,starPoints,totalExpPoints
    ])

wb.save("PlayerProfile.xlsx")

#Calculate end time
end_time = time.time()
print(f"Time：{end_time - start_time}")
