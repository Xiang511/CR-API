import requests
import openpyxl
import datetime
import time
from tqdm import tqdm
from apscheduler.schedulers.background import BackgroundScheduler

scheduler = BackgroundScheduler()

def job_function():
    #Calculate starting time
    start_time = time.time()

    API_KEY = ""
    headers = {
    "Authorization": "Bearer {}".format(API_KEY)
    }

    response = requests.get(
    "https://api.clashroyale.com/v1/locations/57000228/pathoflegend/players",
    headers=headers,
    )          

    now = datetime.datetime.now()
    now_str = now.strftime("%Y-%m-%d_%H-%M-%S")

    wb = openpyxl.Workbook() 
    ws = wb.active

    ws.cell(row=1, column=1).value = "ID"
    ws.cell(row=1, column=2).value = "Name"
    ws.cell(row=1, column=3).value = "Rating"

    row_number = 2
    for player in tqdm(response.json()["items"]):
        ws.cell(row=row_number, column=1).value = player["tag"]
        ws.cell(row=row_number, column=2).value = player["name"]
        ws.cell(row=row_number, column=3).value = player["eloRating"]
        row_number += 1

    wb.save(now_str+".xlsx")

    end_time = time.time()
    print(f"Time：{end_time - start_time}")

job_function();


scheduler.start()
scheduler.add_job(job_function, 'interval', seconds=7200)

while(1):
       time.sleep(7200)