import openpyxl
import requests
import json
import time
from tqdm import tqdm

#Calculate starting time
start_time = time.time()

# Enter your key
# Go to https://developer.clashroyale.com/#/ 
API_KEY = ""

headers = {
    "Authorization": "Bearer {}".format(API_KEY)
}

# enter your tag
# ex: '%2322R920J00','%2312345678'

player_tags = ["%2322R920J00", "%23V99082R9C"]



wb = openpyxl.Workbook()
ws = wb.active

ws["A1"] = "Tag"
ws["B1"] = "BEST"
ws["C1"] = "LAST"


def clean_result(result):
    if result is None:
        return ""
    return result[result.rfind(" ") + 1:].split("}")[0]
#ex:
#   best = json.dumps(player_data.get("bestPathOfLegendSeasonResult"))
#   回傳值為{"leagueNumber": 10, "trophies": 2416, "rank": 2573}
#
#   另 best2 = clean_result(best)
#   則 best2 = 2573

for player_tag in tqdm(player_tags):

    response = requests.get(f"https://api.clashroyale.com/v1/players/{player_tag}", headers=headers)
    
    player_data = response.json()

    best = json.dumps(player_data.get("bestPathOfLegendSeasonResult"))
    last = json.dumps(player_data.get("lastPathOfLegendSeasonResult"))

    best2 = clean_result(best)
    last2= clean_result(last)

    ws.append([
    player_data["name"],
    best2,
    last2])

wb.save("Player.xlsx")

#Calculate end time
end_time = time.time()
print(f"Time：{end_time - start_time}")
